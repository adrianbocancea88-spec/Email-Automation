from __future__ import annotations

import os
import smtplib
import ssl
import sys
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from numbers import Real
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - handled at runtime for clearer errors.
    load_workbook = None
    get_column_letter = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None


EXCEL_MIME_TYPE = (
    "application",
    "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
BASE_DIR = Path(__file__).resolve().parent


class EmailReportError(Exception):
    """Base exception for expected application errors."""


class ConfigError(EmailReportError):
    """Raised when required configuration is missing or invalid."""


class ReportReadError(EmailReportError):
    """Raised when the Excel report cannot be read."""


class EmailSendError(EmailReportError):
    """Raised when the email cannot be sent."""


@dataclass(frozen=True)
class EmailConfig:
    report_path: Path
    log_path: Path
    smtp_host: str
    smtp_port: int
    smtp_security: str
    smtp_username: str
    smtp_password: str
    sender: str
    recipients: list[str]
    subject: str
    body_intro: str


@dataclass
class ColumnStats:
    name: str
    count: int = 0
    total: float = 0.0
    minimum: float | None = None
    maximum: float | None = None

    def add(self, value: float) -> None:
        self.count += 1
        self.total += value
        self.minimum = value if self.minimum is None else min(self.minimum, value)
        self.maximum = value if self.maximum is None else max(self.maximum, value)

    @property
    def average(self) -> float | None:
        if self.count == 0:
            return None
        return self.total / self.count


def parse_recipients(raw_recipients: str) -> list[str]:
    recipients = [
        email.strip()
        for email in raw_recipients.replace(";", ",").split(",")
        if email.strip()
    ]
    if not recipients:
        raise ConfigError("EMAIL_TO must include at least one recipient.")
    return recipients


def parse_smtp_port(raw_port: str) -> int:
    try:
        port = int(raw_port)
    except ValueError as exc:
        raise ConfigError("SMTP_PORT must be a valid number.") from exc

    if not 1 <= port <= 65535:
        raise ConfigError("SMTP_PORT must be between 1 and 65535.")
    return port


def resolve_app_path(raw_path: str) -> Path:
    path = Path(raw_path).expanduser()
    if path.is_absolute():
        return path
    return BASE_DIR / path


def load_config() -> EmailConfig:
    smtp_host = os.getenv("SMTP_HOST", "").strip()
    sender = os.getenv("EMAIL_FROM", "").strip()
    recipients_raw = os.getenv("EMAIL_TO", "").strip()

    missing = [
        name
        for name, value in {
            "SMTP_HOST": smtp_host,
            "EMAIL_FROM": sender,
            "EMAIL_TO": recipients_raw,
        }.items()
        if not value
    ]
    if missing:
        raise ConfigError(f"Missing required environment variables: {', '.join(missing)}")

    smtp_username = os.getenv("SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip()
    if bool(smtp_username) != bool(smtp_password):
        raise ConfigError(
            "Set both SMTP_USERNAME and SMTP_PASSWORD, or leave both empty."
        )

    smtp_security = os.getenv("SMTP_SECURITY", "starttls").strip().lower()
    if smtp_security not in {"starttls", "ssl", "none"}:
        raise ConfigError("SMTP_SECURITY must be one of: starttls, ssl, none.")

    body_intro = os.getenv(
        "EMAIL_BODY",
        "Hello,\\n\\nPlease find attached the latest Excel report.",
    ).replace("\\n", "\n")

    return EmailConfig(
        report_path=resolve_app_path(os.getenv("REPORT_FILE") or "report.xlsx"),
        log_path=resolve_app_path(os.getenv("LOG_FILE") or "log.txt"),
        smtp_host=smtp_host,
        smtp_port=parse_smtp_port(os.getenv("SMTP_PORT", "587")),
        smtp_security=smtp_security,
        smtp_username=smtp_username,
        smtp_password=smtp_password,
        sender=sender,
        recipients=parse_recipients(recipients_raw),
        subject=os.getenv("EMAIL_SUBJECT", "Automated Excel Report").strip()
        or "Automated Excel Report",
        body_intro=body_intro,
    )


def format_number(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value:,.2f}".rstrip("0").rstrip(".")


def normalize_headers(raw_headers: Iterable[object]) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(raw_headers, start=1):
        if value is None or str(value).strip() == "":
            column_letter = get_column_letter(index) if get_column_letter else str(index)
            headers.append(f"Column {column_letter}")
        else:
            headers.append(str(value).strip())
    return headers


def is_numeric(value: object) -> bool:
    return isinstance(value, Real) and not isinstance(value, bool)


def build_sheet_summary(sheet) -> str:
    rows = sheet.iter_rows(values_only=True)

    header_row = None
    for row in rows:
        if any(value is not None for value in row):
            header_row = row
            break

    if header_row is None:
        return f"Sheet: {sheet.title}\n- Status: empty"

    headers = normalize_headers(header_row)
    numeric_stats = [ColumnStats(name=header) for header in headers]
    data_rows = 0

    for row in rows:
        if all(value is None for value in row):
            continue

        data_rows += 1
        for index, value in enumerate(row[: len(headers)]):
            if is_numeric(value):
                numeric_stats[index].add(float(value))

    populated_numeric_stats = [stats for stats in numeric_stats if stats.count > 0]

    lines = [
        f"Sheet: {sheet.title}",
        f"- Data rows: {data_rows}",
        f"- Columns: {len(headers)}",
    ]

    if not populated_numeric_stats:
        lines.append("- Numeric columns: none found")
        return "\n".join(lines)

    lines.append("- Numeric columns:")
    for stats in populated_numeric_stats:
        lines.append(
            "  * "
            f"{stats.name}: count={stats.count}, "
            f"total={format_number(stats.total)}, "
            f"avg={format_number(stats.average)}, "
            f"min={format_number(stats.minimum)}, "
            f"max={format_number(stats.maximum)}"
        )

    return "\n".join(lines)


def build_report_summary(report_path: Path) -> str:
    if load_workbook is None:
        raise ReportReadError(
            "The 'openpyxl' package is not installed. "
            "Run 'pip install -r requirements.txt' and try again."
        ) from OPENPYXL_IMPORT_ERROR

    if not report_path.is_file():
        raise ReportReadError(f"Report file not found: {report_path}")

    try:
        workbook = load_workbook(report_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ReportReadError(f"Could not open report file '{report_path}': {exc}") from exc

    try:
        summaries = [build_sheet_summary(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    return "\n\n".join(summaries)


def build_email_body(body_intro: str, summary: str) -> str:
    return (
        f"{body_intro.strip()}\n\n"
        "Report summary:\n\n"
        f"{summary}\n\n"
        "Regards,\n"
        "Automated Email Reporter"
    )


def create_email_message(config: EmailConfig, body: str) -> EmailMessage:
    message = EmailMessage()
    message["From"] = config.sender
    message["To"] = ", ".join(config.recipients)
    message["Subject"] = config.subject
    message.set_content(body)

    try:
        attachment_data = config.report_path.read_bytes()
    except OSError as exc:
        raise ReportReadError(
            f"Could not read attachment '{config.report_path}': {exc}"
        ) from exc

    maintype, subtype = EXCEL_MIME_TYPE
    message.add_attachment(
        attachment_data,
        maintype=maintype,
        subtype=subtype,
        filename=config.report_path.name,
    )
    return message


def send_email(config: EmailConfig, message: EmailMessage) -> None:
    context = ssl.create_default_context()

    try:
        if config.smtp_security == "ssl":
            with smtplib.SMTP_SSL(
                config.smtp_host,
                config.smtp_port,
                timeout=30,
                context=context,
            ) as smtp:
                authenticate_and_send(smtp, config, message)
        else:
            with smtplib.SMTP(config.smtp_host, config.smtp_port, timeout=30) as smtp:
                smtp.ehlo()
                if config.smtp_security == "starttls":
                    smtp.starttls(context=context)
                    smtp.ehlo()
                authenticate_and_send(smtp, config, message)
    except (OSError, smtplib.SMTPException) as exc:
        raise EmailSendError(f"Could not send email: {exc}") from exc


def authenticate_and_send(
    smtp: smtplib.SMTP,
    config: EmailConfig,
    message: EmailMessage,
) -> None:
    if config.smtp_username and config.smtp_password:
        smtp.login(config.smtp_username, config.smtp_password)
    smtp.send_message(message)


def current_timestamp() -> str:
    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def append_log(log_path: Path, line: str) -> None:
    if log_path.parent != Path("."):
        log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as log_file:
        log_file.write(line + "\n")


def log_sent_email(config: EmailConfig) -> None:
    append_log(
        config.log_path,
        (
            f"{current_timestamp()} | SENT | "
            f"To: {', '.join(config.recipients)} | "
            f"Subject: {config.subject} | "
            f"Attachment: {config.report_path}"
        ),
    )


def log_error(log_path: Path, error: Exception) -> None:
    append_log(log_path, f"{current_timestamp()} | ERROR | {error}")


def safe_log_error(log_path: Path, error: Exception) -> None:
    try:
        log_error(log_path, error)
    except Exception as log_exc:
        print(
            f"Warning: could not write to log file '{log_path}': {log_exc}",
            file=sys.stderr,
        )


def main() -> int:
    log_path = resolve_app_path(os.getenv("LOG_FILE") or "log.txt")

    try:
        config = load_config()
        summary = build_report_summary(config.report_path)
        body = build_email_body(config.body_intro, summary)
        message = create_email_message(config, body)
        send_email(config, message)
        log_sent_email(config)
    except EmailReportError as exc:
        safe_log_error(log_path, exc)
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        safe_log_error(log_path, exc)
        print(f"Unexpected error: {exc}", file=sys.stderr)
        return 1

    print(f"Email sent successfully to {', '.join(config.recipients)}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
